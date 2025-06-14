import os
import re
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font


def remove_paragraph_indents(text):
    cleaned_lines = [line.lstrip() for line in text.splitlines()]
    return '\n'.join(cleaned_lines)


stop_words = {
    'کجا', 'نبودی', 'بودی', 'شدی', 'nn', 'آنجا', 'آنکه', 'های', 'ست', 'کرده',
    'ازین', 'ازان', 'باز', 'گردد', '،', 'گشت', 'فی', 'زین', 'کو', 'هست', 'چنین',
    'برای', 'شود', 'کس', 'کز', 'بهر', 'همچو', 'سوی', 'پی', 'همی', 'شمارهٔ', 'یا',
    'چو', 'مر', 'گفت', 'برات', 'چون', 'هیچ', 'صد', 'دارد', 'داری', 'نیست', 'است',
    'گوید', 'چی', 'چه', 'اگر', 'هر', 'را', 'از', 'ز', 'چهار', 'سه', 'یک', 'شما',
    'وی', 'انها', 'ما', 'او', 'تو', 'من', 'دو', 'پیش', 'گر', 'می', 'و', 'در', 'کی',
    'شد', 'کرد', 'گرچه', 'گاه', 'گه', 'زان', 'که', 'به', 'بر', 'آن', 'تا', 'این',
    'اندر', 'آندر', 'اینجا', 'درین', 'بود', 'بی', 'آمد', 'چون', 'با', 'ای', 'هم',
    'باشد', 'ترا', '-', ':', 'سر', 'شماره', 'شده', 'شد', 'ام', 'وز'
}

keywords = ['علم', 'عمل', 'معرفت', 'یقین']

folder_path = r'C:\Users\Lenovo\Documents\корпус персидской поэзии\attar'
texts = []


def clean_text(text, keep_words=None):

    if keep_words is None:
        keep_words = set()

    tokens = re.split(r'(\s+)', text)
    cleaned_tokens = []

    for token in tokens:
        if token.strip() in keep_words:
            cleaned_tokens.append(token)
        elif token.isspace():
            cleaned_tokens.append(token)
        else:
            cleaned_token = re.sub(r'\d+', '', token)  # Удаляем цифры
            cleaned_token = re.sub(r'[^\w\s]', '', cleaned_token)  # Удаляем знаки препинания
            cleaned_tokens.append(cleaned_token)

    return ''.join(cleaned_tokens)


for filename in os.listdir(folder_path):
    if filename.endswith('.txt'):
        with open(os.path.join(folder_path, filename), 'r', encoding='utf-8') as file:
            content = file.read()
            content = clean_text(content, set(keywords))
            cleaned_content = remove_paragraph_indents(content)
            texts.append(cleaned_content)


def filter_words(text, stop_words, keywords):
    words = text.split()
    filtered_words = []

    for word in words:
        if word in keywords:
            filtered_words.append(word)
        elif word not in stop_words and word.strip() != '':
            filtered_words.append(word)

    return filtered_words


corpus_words = []
for text in texts:
    corpus_words.extend(filter_words(text, stop_words, keywords))


def find_context_words(words_list, keywords, stop_words, window_size=40):
    context_words = {keyword: [] for keyword in keywords}

    for i, word in enumerate(words_list):
        if word in keywords:
            start_index = max(0, i - window_size)
            end_index = min(len(words_list), i + window_size + 1)

            context = [
                w for w in words_list[start_index:end_index]
                if (w != word and (w in keywords or (w not in stop_words and w.strip() != '')))
            ]
            context_words[word].extend(context)

    return context_words


context_words_all = find_context_words(corpus_words, keywords, stop_words)

context_results = {}
for keyword, contexts in context_words_all.items():
    context_results[keyword] = Counter(contexts)


def create_excel_table(results, filename="collocations_table.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Коллокации"

    for col, keyword in enumerate(results.keys(), start=1):
        cell = ws.cell(row=1, column=col, value=keyword)
        cell.font = Font(bold=True)

    max_collocations = max(len(word_counts.most_common()) for word_counts in results.values())

    for row in range(2, max_collocations + 2):
        for col, (keyword, word_counts) in enumerate(results.items(), start=1):
            collocations = word_counts.most_common()
            if row - 2 < len(collocations):
                word, freq = collocations[row - 2]
                cell = ws.cell(row=row, column=col, value=f"{word} ({freq})")

    for col, keyword in enumerate(results.keys(), start=1):
        max_len = max(len(f"{word} ({freq})") for word, freq in results[keyword].most_common()) if results[
            keyword].most_common() else len(keyword)
        ws.column_dimensions[chr(64 + col)].width = max_len + 2

    wb.save(filename)
    return filename


excel_file = create_excel_table(context_results)
print(f"Результаты сохранены в файл: {excel_file}")

for keyword, counts in context_results.items():
    print(f"\nТоп коллокаций для '{keyword}':")
    for word, freq in counts.most_common(15):
        print(f"{word} ({freq})")